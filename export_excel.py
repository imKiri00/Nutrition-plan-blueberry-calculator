import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_excel_report(preparati_values, parcele_values, results, num_runs, filename="izvjestaj.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Borovnica Prehrana Izvještaj"

    # Set column widths
    ws.column_dimensions['A'].width = 30  # Double width for column A
    for col in range(2, 4):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Title
    ws['A1'] = "Borovnica Prehrana - Izvještaj"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:C1')

    # Unesene vrijednosti
    ws['A3'] = "Unesene vrijednosti:"
    ws['A3'].font = Font(bold=True)

    # Preparati
    ws['A4'] = "Preparati:"
    ws['A4'].font = Font(bold=True)
    for i, (preparat, value) in enumerate(preparati_values.items(), start=5):
        ws[f'A{i}'] = preparat
        ws[f'B{i}'] = f"{value:.2f} g/sadnica"

    # Parcele
    ws['A' + str(5 + len(preparati_values))] = "Parcele:"
    ws['A' + str(5 + len(preparati_values))].font = Font(bold=True)
    for i, (parcela, value) in enumerate(parcele_values.items(), start=6+len(preparati_values)):
        ws[f'A{i}'] = parcela
        ws[f'B{i}'] = f"{value} sadnica"

    # Izračunate količine
    row = max(len(preparati_values), len(parcele_values)) + 7
    ws[f'A{row}'] = f"Izračunate količine preparata po parceli (Broj nanošenja: {num_runs}):"
    ws[f'A{row}'].font = Font(bold=True)
    ws.merge_cells(f'A{row}:C{row}')

    # Define styles
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    row += 1
    for parcela, preparat_values in results.items():
        # Parcela header
        ws[f'A{row}'] = f"Parcela {parcela}:"
        ws[f'A{row}'].font = Font(bold=True)
        ws.row_dimensions[row].height = 30  # Double the height
        
        # Merge cells for "O" characters
        ws.merge_cells(f'B{row}:C{row}')
        cell = ws[f'B{row}']
        cell.value = "O" * num_runs
        cell.font = Font(size=24)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Apply gray background to the header row
        for col in range(1, 4):
            ws[f'{get_column_letter(col)}{row}'].fill = header_fill

        row += 1
        start_row = row
        for preparat, value in preparat_values.items():
            total_amount = value
            split_amount = total_amount / num_runs
            ws[f'A{row}'] = preparat
            ws[f'B{row}'] = f"{total_amount:.2f} g"
            ws[f'C{row}'] = f"{split_amount:.2f} g x {num_runs}"
            row += 1
        
        # Apply border to the entire block
        for r in range(start_row - 1, row):
            for col in range(1, 4):
                ws[f'{get_column_letter(col)}{r}'].border = border

        row += 1

    # Apply alignment to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    wb.save(filename)
    return filename

def export_to_excel(preparati_values, parcele_values, results, num_runs):
    return create_excel_report(preparati_values, parcele_values, results, num_runs)